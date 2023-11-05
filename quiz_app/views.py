from django.shortcuts import render, redirect
from .models import Question, Participant, Limit , PDF
from .forms import UserLoginForm
import random
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseRedirect
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from django.contrib import messages
from django.urls import reverse
from django.core.mail import send_mail
from django.conf import settings
from django.contrib.auth import authenticate, login



def index(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)

            # Check if a 'Participant' object already exists for the user
            participant, created = Participant.objects.get_or_create(user=user)

            if created:
                request.session['participant_id'] = participant.id  # Store Participant's id in session
                return redirect('quiz')  # Redirect to the quiz page after login
            else:
                messages.error(request, 'Already attended.')  # Participant already exists
        else:
            messages.error(request, 'Invalid username or password. Please try again.')

    # If the request is not POST, render the sign-in page
    return render(request, 'quiz_app/index.html')
    


def quiz(request):
    if 'participant_id' not in request.session:
        return redirect('index')

    if request.method == 'POST':
        selected_option = request.POST.get('answer')
        correct_option = request.session['correct_answer']
        correct_option2 = request.session['correct_answer2']
        correct_option3 = request.session['correct_answer3']
        correct_option4 = request.session['correct_answer4']
        if selected_option:
            if (selected_option == correct_option
                or selected_option == correct_option2 
                or selected_option == correct_option3 
                or selected_option == correct_option4):
                
                participant = Participant.objects.get(id=request.session['participant_id'])
                participant.score += 1
                participant.save()
            
            # Get the current participant
            participant = Participant.objects.get(id=request.session['participant_id'])

            # Get the current question
            current_question_id = request.session.get('current_question_id')
            current_question = Question.objects.get(id=current_question_id)

            # Update the answers field in participant's model
            answers = participant.answers or {}
            answers[current_question.id] = selected_option
            participant.answers = answers
            participant.save()

        question_count = request.session.get('question_count', 0)
        if question_count >= 29:
            return redirect('result')
        else:
            request.session['question_count'] = question_count + 1

    questions = list(Question.objects.exclude(id__in=request.session.get('seen_question_ids', [])))
    if not questions:
        return redirect('result')

    random_question = random.choice(questions)
    request.session['correct_answer'] = random_question.correct_answer
    request.session['correct_answer2'] = random_question.correct_answer2
    request.session['correct_answer3'] = random_question.correct_answer3
    request.session['correct_answer4'] = random_question.correct_answer4

    request.session['current_question_id'] = random_question.id

    seen_question_ids = request.session.get('seen_question_ids', [])
    seen_question_ids.append(random_question.id)
    request.session['seen_question_ids'] = seen_question_ids

    return render(request, 'quiz_app/quiz.html', {'question': random_question, 'question_count': request.session.get('question_count', 0) + 1, 'total_questions': 30})


def result(request):
    if 'participant_id' not in request.session:
        return redirect('index')

    participant = Participant.objects.get(id=request.session['participant_id'])
    
    quiz = Question.objects.all()
    
    context = {
        'score': participant.score,
        'won_gift': participant.score >= 15,
        'answers' : participant.answers,
        'quiz' : quiz
    }


#    if participant.score >= 7:
#        subject = "Price has won"
#        message = f"\n\nName: {participant.name}\nScore: {participant.score}\nPhone: {participant.phone_number}"
#        from_email = settings.DEFAULT_FROM_EMAIL
#        recipient_list = ["sreyaslove@gmail.com","podapattee007@gmail.com","saravanansvanas426@gmail.com"]
#        send_mail(subject, message, from_email, recipient_list)
#    else:
#        subject = "Sorry, he lost"
#        message = f"\n\nName: {participant.name}\nScore: {participant.score}\nPhone: {participant.phone_number}"
#        from_email = settings.DEFAULT_FROM_EMAIL
#        recipient_list = ["sreyaslove@gmail.com","podapattee007@gmail.com","saravanansvanas426@gmail.com"]
#        send_mail(subject, message, from_email, recipient_list)



    request.session.flush()
    return render(request, 'quiz_app/result.html', context)



def export_participants_xlsx(request):
    participants = Participant.objects.all()

    # Create a workbook and add a worksheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Participants'

    # Add headers to the worksheet
    headers = ['Name', 'Phone Number', 'Score']
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    # Add participant data to the worksheet
    for row_num, participant in enumerate(participants, 2):
        sheet.cell(row=row_num, column=1, value=participant.name)
        sheet.cell(row=row_num, column=2, value=participant.phone_number)
        sheet.cell(row=row_num, column=3, value=participant.score)

    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=participants.xlsx'
    workbook.save(response)

    return response


def export_participants_pdf(request):
    participants = Participant.objects.all()

    # Create a PDF document
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=participants.pdf'
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []

    # Add participant data to the table
    data = [['Name', 'Phone Number', 'Score']]
    for participant in participants:
        data.append([participant.name, participant.phone_number, participant.score])

    table = Table(data)

    # Apply table styles
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 1), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.beige]),
        ('ROWBACKGROUNDS', (0, 0), (-1, 0), [colors.grey])
    ]))

    elements.append(table)
    doc.build(elements)

    return response

@login_required
def dashboard(request):
    return render(request,'quiz_app/admin_dashboard.html')
    
    
    
    
def study(request):
    
    pdf = PDF.objects.all()
    
    print(PDF)
    
    return render(request,'quiz_app/studymaterial.html' , {'pdf':pdf})
