from django.http import HttpResponse
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

def export_participants_xlsx(modeladmin, request, queryset):
    # Create a workbook and add a worksheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Participants'

    # Add headers to the worksheet
    headers = ['Name', 'Phone Number', 'Score']
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    # Add participant data to the worksheet
    for row_num, participant in enumerate(queryset, 2):
        sheet.cell(row=row_num, column=1, value=participant.name)
        sheet.cell(row=row_num, column=2, value=participant.phone_number)
        sheet.cell(row=row_num, column=3, value=participant.score)

    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=participants.xlsx'
    workbook.save(response)

    return response
export_participants_xlsx.short_description = "Export selected participants to Excel"

def export_participants_pdf(modeladmin, request, queryset):
    # Create a PDF document
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=participants.pdf'
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []

    # Add participant data to the table
    data = [['Name', 'Phone Number', 'Score']]
    for participant in queryset:
        data.append([participant.name, participant.phone_number, participant.score])

    table = Table(data)

    # Apply table styles
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 1), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.beige]),
        ('ROWBACKGROUNDS', (0, 0), (-1, 0), [colors.grey])
    ]))

    elements.append(table)
    doc.build(elements)

    return response
    
export_participants_pdf.short_description = "Export selected participants to PDF"
